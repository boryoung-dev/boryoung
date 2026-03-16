"""상품등록 엑셀 템플릿 생성 스크립트"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ============================================
# Sheet 1: 상품 기본정보
# ============================================
ws = wb.active
ws.title = "상품 기본정보"

# 스타일 정의
header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
required_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
required_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
example_font = Font(name="맑은 고딕", size=10, color="666666")
example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
note_font = Font(name="맑은 고딕", size=9, color="FF0000", italic=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 컬럼 정의: (헤더명, 너비, 필수여부, 설명)
columns = [
    ("상품명", 35, True, "상품 제목 (예: 베트남 다낭 3박5일 골프투어)"),
    ("부제목", 30, False, "상품 부제목 (예: 프리미엄 3색 라운딩)"),
    ("카테고리", 15, True, "동남아골프 / 일본골프 / 국내골프 / 제주골프 등"),
    ("목적지", 15, True, "여행 목적지 (예: 베트남 다낭, 태국 파타야)"),
    ("출발지", 12, False, "인천 / 김포 / 부산 / 대구 등"),
    ("항공사", 15, False, "항공사명 (예: 대한항공, 비엣젯항공)"),
    ("숙박(박)", 10, True, "숙박 일수 (숫자만)"),
    ("여행(일)", 10, True, "여행 일수 (숫자만)"),
    ("기간 텍스트", 15, False, "표시용 (예: 3박5일). 비우면 자동생성"),
    ("골프코스명", 25, False, "골프코스 이름 (여러개면 쉼표로 구분)"),
    ("총 홀수", 10, False, "총 라운딩 홀 수 (숫자만, 예: 54)"),
    ("난이도", 12, False, "초급 / 중급 / 상급 / 전체"),
    ("최소인원", 10, False, "최소 출발 인원 (숫자만)"),
    ("최대인원", 10, False, "최대 인원 (숫자만)"),
    ("판매가(원)", 15, True, "1인 기준 판매가 (숫자만, 예: 1290000)"),
    ("정가(원)", 15, False, "할인 전 정가 (숫자만). 할인 표시용"),
    ("포함사항", 40, False, "쉼표로 구분 (예: 왕복항공료, 숙박, 그린피, 카트비)"),
    ("불포함사항", 40, False, "쉼표로 구분 (예: 여행자보험, 개인경비, 캐디피)"),
    ("상품설명", 50, False, "상품 상세 설명"),
    ("출발일정", 30, False, "출발 가능일 (쉼표로 구분, 예: 2026-03-15, 2026-03-22)"),
    ("추천상품", 10, False, "Y / N (메인 노출 여부)"),
    ("네이버 URL", 35, False, "네이버 스마트스토어 등 외부 링크"),
]

# 안내 행 (1행)
ws.merge_cells("A1:V1")
ws["A1"] = "※ 빨간색 헤더 = 필수입력 | 파란색 헤더 = 선택입력 | 회색 행 = 예시 데이터 (삭제 후 입력해주세요)"
ws["A1"].font = Font(name="맑은 고딕", size=10, bold=True, color="C00000")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 25

# 설명 행 (2행)
for col_idx, (name, width, required, desc) in enumerate(columns, 1):
    cell = ws.cell(row=2, column=col_idx, value=desc)
    cell.font = Font(name="맑은 고딕", size=8, color="808080", italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[2].height = 35

# 헤더 행 (3행)
for col_idx, (name, width, required, desc) in enumerate(columns, 1):
    cell = ws.cell(row=3, column=col_idx, value=f"★ {name}" if required else name)
    cell.font = required_font if required else header_font
    cell.fill = required_fill if required else header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[3].height = 30

# 예시 데이터 1
example1 = [
    "베트남 다낭 3박5일 프리미엄 골프투어",
    "3색 명문코스 라운딩 + 5성급 리조트",
    "동남아골프",
    "베트남 다낭",
    "인천",
    "대한항공",
    3,
    5,
    "3박5일",
    "바나힐스 GC, 몽고메리링크스, BRG다낭",
    54,
    "전체",
    4,
    16,
    1290000,
    1590000,
    "왕복항공료, 5성급 숙박(3박), 그린피(3회), 카트비, 전용차량, 조식3회",
    "여행자보험, 개인경비, 캐디피(현지지불), 중석식",
    "베트남 다낭의 명문 3개 코스에서 라운딩하는 프리미엄 골프투어입니다. 5성급 리조트에서 편안한 휴식까지!",
    "2026-03-15, 2026-03-22, 2026-04-05",
    "Y",
    "",
]

# 예시 데이터 2
example2 = [
    "태국 파타야 2박3일 골프패키지",
    "시암CC 포함 2색 라운딩",
    "동남아골프",
    "태국 파타야",
    "인천",
    "타이에어아시아X",
    2,
    3,
    "2박3일",
    "시암CC 올드코스, 라엠차방 인터내셔널CC",
    36,
    "중급",
    2,
    12,
    890000,
    1100000,
    "왕복항공료, 숙박(2박), 그린피(2회), 카트비, 공항픽업",
    "여행자보험, 개인경비, 캐디피(현지지불), 식사비용",
    "태국 파타야에서 즐기는 가성비 골프여행! 시암CC 올드코스 라운딩 포함.",
    "2026-04-10, 2026-04-17",
    "N",
    "https://smartstore.naver.com/example/products/123",
]

for col_idx, value in enumerate(example1, 1):
    cell = ws.cell(row=4, column=col_idx, value=value)
    cell.font = example_font
    cell.fill = example_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border

for col_idx, value in enumerate(example2, 1):
    cell = ws.cell(row=5, column=col_idx, value=value)
    cell.font = example_font
    cell.fill = example_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border

ws.row_dimensions[4].height = 45
ws.row_dimensions[5].height = 45

# 빈 입력 행 (6~25행) 테두리
for row in range(6, 26):
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 25

# 데이터 유효성 검사
# 카테고리 드롭다운
cat_validation = DataValidation(
    type="list",
    formula1='"동남아골프,일본골프,국내골프,제주골프,호주골프,미국골프,유럽골프"',
    allow_blank=True,
)
cat_validation.error = "목록에서 선택해주세요"
cat_validation.errorTitle = "카테고리 오류"
ws.add_data_validation(cat_validation)
cat_validation.add(f"C4:C25")

# 난이도 드롭다운
diff_validation = DataValidation(
    type="list",
    formula1='"초급,중급,상급,전체"',
    allow_blank=True,
)
ws.add_data_validation(diff_validation)
diff_validation.add(f"L4:L25")

# 추천상품 드롭다운
featured_validation = DataValidation(
    type="list",
    formula1='"Y,N"',
    allow_blank=True,
)
ws.add_data_validation(featured_validation)
featured_validation.add(f"U4:U25")

# 시트 고정 (헤더 행)
ws.freeze_panes = "A4"

# ============================================
# Sheet 2: 일정표 (Itinerary)
# ============================================
ws2 = wb.create_sheet("일별 일정")

itin_columns = [
    ("상품명", 35, True, "Sheet1의 상품명과 동일하게 입력"),
    ("일차", 8, True, "숫자만 (예: 1, 2, 3)"),
    ("일정 제목", 25, True, "해당 일차 제목 (예: 인천출발 → 다낭도착)"),
    ("일정 상세", 50, False, "상세 설명"),
    ("식사", 30, False, "예: 조식-호텔식 / 중식-현지식 / 석식-자유식"),
    ("숙소", 25, False, "숙소명 (예: 풀만 다낭 비치리조트)"),
    ("골프코스", 20, False, "해당일 라운딩 코스명"),
    ("라운딩 홀수", 10, False, "해당일 홀 수 (숫자만, 예: 18)"),
    ("이동수단", 15, False, "전용차량 / 택시 / 도보 등"),
]

# 안내 행
ws2.merge_cells("A1:I1")
ws2["A1"] = "※ 상품별 일차를 1일차부터 순서대로 입력해주세요. 상품명은 '상품 기본정보' 시트와 동일하게 입력합니다."
ws2["A1"].font = Font(name="맑은 고딕", size=10, bold=True, color="C00000")
ws2.row_dimensions[1].height = 25

# 설명 행
for col_idx, (name, width, required, desc) in enumerate(itin_columns, 1):
    cell = ws2.cell(row=2, column=col_idx, value=desc)
    cell.font = Font(name="맑은 고딕", size=8, color="808080", italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws2.row_dimensions[2].height = 35

# 헤더
for col_idx, (name, width, required, desc) in enumerate(itin_columns, 1):
    cell = ws2.cell(row=3, column=col_idx, value=f"★ {name}" if required else name)
    cell.font = required_font if required else header_font
    cell.fill = required_fill if required else header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[3].height = 30

# 예시 일정 (상품1)
itinerary_examples = [
    ["베트남 다낭 3박5일 프리미엄 골프투어", 1, "인천 출발 → 다낭 도착", "인천국제공항 출발, 다낭 국제공항 도착 후 호텔 체크인", "석식-환영만찬(현지식)", "풀만 다낭 비치리조트", "", "", "전용차량"],
    ["베트남 다낭 3박5일 프리미엄 골프투어", 2, "바나힐스 GC 라운딩", "오전 바나힐스 골프클럽 18홀 라운딩, 오후 자유시간", "조식-호텔식 / 중식-클럽하우스 / 석식-자유식", "풀만 다낭 비치리조트", "바나힐스 GC", 18, "전용차량"],
    ["베트남 다낭 3박5일 프리미엄 골프투어", 3, "몽고메리링크스 라운딩", "오전 몽고메리링크스 18홀 라운딩, 오후 호이안 관광", "조식-호텔식 / 중식-클럽하우스 / 석식-호이안 현지식", "풀만 다낭 비치리조트", "몽고메리링크스", 18, "전용차량"],
    ["베트남 다낭 3박5일 프리미엄 골프투어", 4, "BRG다낭 라운딩 + 출발", "오전 BRG다낭 18홀 라운딩, 호텔 체크아웃, 공항 이동", "조식-호텔식 / 중식-클럽하우스", "", "BRG다낭", 18, "전용차량"],
    ["베트남 다낭 3박5일 프리미엄 골프투어", 5, "인천 도착", "인천국제공항 도착, 해산", "기내식", "", "", "", ""],
    ["태국 파타야 2박3일 골프패키지", 1, "인천 출발 → 파타야 도착", "인천공항 출발, 방콕 수완나품 도착 후 파타야 이동", "석식-현지식", "아마리 파타야", "", "", "전용차량"],
    ["태국 파타야 2박3일 골프패키지", 2, "시암CC + 라엠차방CC 라운딩", "오전 시암CC 올드코스 18홀, 오후 라엠차방CC 18홀", "조식-호텔식 / 중식-클럽하우스 / 석식-현지식", "아마리 파타야", "시암CC 올드코스, 라엠차방CC", 36, "전용차량"],
    ["태국 파타야 2박3일 골프패키지", 3, "파타야 → 인천 도착", "호텔 체크아웃, 방콕 공항 이동, 인천 도착", "조식-호텔식", "", "", "", "전용차량"],
]

for row_idx, data in enumerate(itinerary_examples, 4):
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = example_font
        cell.fill = example_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[row_idx].height = 35

# 빈 입력 행
for row in range(12, 52):
    for col_idx in range(1, len(itin_columns) + 1):
        cell = ws2.cell(row=row, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

ws2.freeze_panes = "A4"

# ============================================
# Sheet 3: 가격옵션
# ============================================
ws3 = wb.create_sheet("가격 옵션")

price_columns = [
    ("상품명", 35, True, "Sheet1의 상품명과 동일하게 입력"),
    ("옵션명", 20, True, "예: 성인, 아동, 싱글룸 추가 등"),
    ("옵션 설명", 30, False, "옵션에 대한 설명"),
    ("가격(원)", 15, True, "숫자만 입력"),
    ("가격유형", 12, True, "1인당 / 1실당 / 추가비용"),
    ("시즌", 10, False, "성수기 / 비수기 / 일반"),
    ("적용시작일", 15, False, "YYYY-MM-DD 형식"),
    ("적용종료일", 15, False, "YYYY-MM-DD 형식"),
    ("기본옵션", 10, False, "Y / N (대표 가격 여부)"),
]

ws3.merge_cells("A1:I1")
ws3["A1"] = "※ 상품별 가격 옵션을 입력합니다. 기본 판매가 외에 추가 옵션이 있을 때 사용합니다."
ws3["A1"].font = Font(name="맑은 고딕", size=10, bold=True, color="C00000")
ws3.row_dimensions[1].height = 25

for col_idx, (name, width, required, desc) in enumerate(price_columns, 1):
    cell = ws3.cell(row=2, column=col_idx, value=desc)
    cell.font = Font(name="맑은 고딕", size=8, color="808080", italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws3.row_dimensions[2].height = 35

for col_idx, (name, width, required, desc) in enumerate(price_columns, 1):
    cell = ws3.cell(row=3, column=col_idx, value=f"★ {name}" if required else name)
    cell.font = required_font if required else header_font
    cell.fill = required_fill if required else header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws3.column_dimensions[get_column_letter(col_idx)].width = width
ws3.row_dimensions[3].height = 30

price_examples = [
    ["베트남 다낭 3박5일 프리미엄 골프투어", "성인", "만 12세 이상", 1290000, "1인당", "일반", "2026-03-01", "2026-06-30", "Y"],
    ["베트남 다낭 3박5일 프리미엄 골프투어", "아동", "만 12세 미만 (라운딩 미포함)", 890000, "1인당", "일반", "2026-03-01", "2026-06-30", "N"],
    ["베트남 다낭 3박5일 프리미엄 골프투어", "싱글룸 추가", "1인실 사용 시 추가요금", 250000, "추가비용", "", "", "", "N"],
    ["태국 파타야 2박3일 골프패키지", "성인", "만 12세 이상", 890000, "1인당", "일반", "2026-04-01", "2026-06-30", "Y"],
    ["태국 파타야 2박3일 골프패키지", "성수기 성인", "만 12세 이상 (7~8월)", 1050000, "1인당", "성수기", "2026-07-01", "2026-08-31", "N"],
]

for row_idx, data in enumerate(price_examples, 4):
    for col_idx, value in enumerate(data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.font = example_font
        cell.fill = example_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
    ws3.row_dimensions[row_idx].height = 30

# 드롭다운
price_type_val = DataValidation(type="list", formula1='"1인당,1실당,추가비용"', allow_blank=True)
ws3.add_data_validation(price_type_val)
price_type_val.add("E4:E30")

season_val = DataValidation(type="list", formula1='"성수기,비수기,일반"', allow_blank=True)
ws3.add_data_validation(season_val)
season_val.add("F4:F30")

default_val = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws3.add_data_validation(default_val)
default_val.add("I4:I30")

for row in range(9, 31):
    for col_idx in range(1, len(price_columns) + 1):
        cell = ws3.cell(row=row, column=col_idx)
        cell.border = thin_border
ws3.freeze_panes = "A4"

# ============================================
# Sheet 4: 작성 가이드
# ============================================
ws4 = wb.create_sheet("작성 가이드")
ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 60
ws4.column_dimensions["C"].width = 40

guide_header = Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
guide_body = Font(name="맑은 고딕", size=10)
guide_bold = Font(name="맑은 고딕", size=10, bold=True)

ws4["A1"] = "상품등록 엑셀 작성 가이드"
ws4["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="2F5496")
ws4.merge_cells("A1:C1")
ws4.row_dimensions[1].height = 35

guides = [
    ("", "", ""),
    ("시트 구성", "설명", "비고"),
    ("상품 기본정보", "상품의 기본 정보를 입력합니다 (1상품 = 1행)", "★ 표시 = 필수입력"),
    ("일별 일정", "상품별 일차별 일정을 입력합니다", "상품명으로 매칭"),
    ("가격 옵션", "성인/아동/시즌별 가격을 입력합니다", "기본 판매가 외 추가옵션"),
    ("", "", ""),
    ("입력 규칙", "설명", "예시"),
    ("상품명", "모든 시트에서 동일한 상품명 사용 필수", "베트남 다낭 3박5일 프리미엄 골프투어"),
    ("숫자 필드", "숫자만 입력 (콤마, 원 등 제외)", "1290000 (O) / 1,290,000원 (X)"),
    ("날짜", "YYYY-MM-DD 형식", "2026-03-15"),
    ("다중값", "쉼표(,)로 구분하여 입력", "왕복항공료, 숙박, 그린피"),
    ("Y/N 필드", "Y 또는 N만 입력", "추천상품: Y"),
    ("", "", ""),
    ("카테고리 목록", "사용 가능한 카테고리", ""),
    ("", "동남아골프, 일본골프, 국내골프, 제주골프, 호주골프, 미국골프, 유럽골프", "드롭다운 선택 가능"),
    ("", "", ""),
    ("난이도 목록", "사용 가능한 난이도", ""),
    ("", "초급, 중급, 상급, 전체", "드롭다운 선택 가능"),
]

for row_idx, (a, b, c) in enumerate(guides, 2):
    ws4.cell(row=row_idx, column=1, value=a).font = guide_bold if a else guide_body
    ws4.cell(row=row_idx, column=2, value=b).font = guide_body
    ws4.cell(row=row_idx, column=3, value=c).font = guide_body

# 시트 구성/입력 규칙 헤더 행 강조
for row in [3, 8, 15, 18]:
    for col in range(1, 4):
        cell = ws4.cell(row=row, column=col)
        if cell.value:
            cell.font = Font(name="맑은 고딕", bold=True, size=10, color="2F5496")

# 저장
output_path = "/Users/simjaehyeong/Desktop/side/boryoung/상품등록_템플릿.xlsx"
wb.save(output_path)
print(f"엑셀 템플릿 생성 완료: {output_path}")
